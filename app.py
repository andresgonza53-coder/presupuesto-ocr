
import io
import re
from pathlib import Path
from datetime import datetime

import cv2
import fitz
import numpy as np
import pandas as pd
import pytesseract
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Presupuestos OCR",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded",
)

APP_NAME = "Presupuestos OCR"
APP_VERSION = "2.1"

STANDARD_COLUMNS = [
    "Item", "Código", "Descripción", "Referencia", "Marca",
    "Cantidad", "Precio Unitario", "Precio Total", "Impuesto", "Plazo"
]

st.markdown("""
<style>
.block-container {max-width: 1500px; padding-top: 2rem;}
.hero-title {font-size: 2.25rem; font-weight: 800; line-height:1.1;}
.hero-subtitle {color:#6b7280; margin:.35rem 0 1.2rem 0;}
.provider {
  display:inline-block; padding:.25rem .65rem; border-radius:999px;
  background:#ecfdf5; border:1px solid #a7f3d0; font-weight:600;
}
div[data-testid="stMetric"] {
  border:1px solid #e5e7eb; padding:.75rem 1rem; border-radius:12px;
}
</style>
""", unsafe_allow_html=True)


# ============================================================
# UTILIDADES
# ============================================================
def clean_text(value):
    value = str(value or "").replace("\x0c", " ")
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r"\s*\n\s*", " ", value)
    return value.strip(" -:\t\r\n")


def clean_code(value):
    return re.sub(r"\s+", "", clean_text(value))


def parse_py_amount(value):
    """Importes paraguayos: 69.345,00 / 693.450 / 1.124.995 / 1,00."""
    s = clean_text(value).upper()
    s = s.replace("GS", "").replace("G$", "")
    s = re.sub(r"[^\d,.\-]", "", s)
    if not s:
        return None

    if "," in s and "." in s:
        # 69.345,00
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        # 1,00
        s = s.replace(",", ".")
    elif "." in s:
        parts = s.split(".")
        if len(parts[-1]) == 3:
            s = "".join(parts)

    try:
        return float(s)
    except Exception:
        return None


def format_py_amount(value):
    if value is None or value == "":
        return ""
    try:
        n = float(value)
        if abs(n - int(n)) < 1e-9:
            return f"{int(n):,}".replace(",", ".")
        return f"{n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(value)


def normalize_numeric_columns(df):
    if df.empty:
        return df
    df = df.copy()
    for col in ["Cantidad", "Precio Unitario", "Precio Total"]:
        if col in df.columns:
            converted = df[col].map(parse_py_amount)
            # Solo convertir si la mayoría de celdas útiles se pudieron interpretar.
            useful = df[col].astype(str).str.strip().ne("").sum()
            ok = converted.notna().sum()
            if useful and ok / useful >= 0.65:
                df[col] = converted
    return df


def empty_standard_df():
    return pd.DataFrame(columns=STANDARD_COLUMNS)


def ensure_standard_columns(df):
    if df is None or df.empty:
        return empty_standard_df()
    df = df.copy()
    for col in STANDARD_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[STANDARD_COLUMNS]


def native_pdf_text(data):
    doc = fitz.open(stream=data, filetype="pdf")
    return "\n".join(page.get_text("text") for page in doc)


def detect_provider(native_text, filename=""):
    t = (native_text or "").lower()
    if "electropar.com.py" in t or ("cotización nº" in t and "código de cliente" in t):
        return "Electropar"
    if "presup. chile nro" in t or "compañía comercial del paraguay" in t:
        return "Compañía Comercial del Paraguay"
    if "electro system" in t or "presupuesto de venta" in t:
        return "Electro System"
    return "Genérico / escaneado"


# ============================================================
# PARSER PDF NATIVO: ELECTROPAR
# ============================================================
def words_by_band(page, y0, y1):
    return [w for w in page.get_text("words") if y0 <= w[1] < y1]


def band_text(words, x0, x1):
    selected = [w for w in words if x0 <= (w[0] + w[2]) / 2 < x1]
    selected.sort(key=lambda w: (round(w[1], 1), w[0]))
    return clean_text(" ".join(w[4] for w in selected))


def parse_electropar_rows(data):
    doc = fitz.open(stream=data, filetype="pdf")
    rows = []

    for page in doc:
        words = page.get_text("words")
        # Inicio de artículos: números de ítem a la izquierda, debajo del encabezado.
        starters = []
        for w in words:
            x0, y0, x1, y1, txt = w[:5]
            if 20 <= x0 < 50 and y0 > 245 and re.fullmatch(r"\d{1,3}", txt):
                starters.append((int(txt), y0))

        starters = sorted(set(starters), key=lambda x: x[1])
        for i, (item, sy) in enumerate(starters):
            ey = starters[i + 1][1] - 2 if i + 1 < len(starters) else min(page.rect.height - 25, sy + 80)
            ws = words_by_band(page, sy - 3, ey)

            row = {
                "Item": item,
                "Cantidad": band_text(ws, 55, 92),
                "Código": band_text(ws, 92, 132),
                "Descripción": band_text(ws, 132, 210),
                "Referencia": clean_code(band_text(ws, 210, 255)),
                "Marca": "",
                "Precio Unitario": band_text(ws, 255, 325).replace("GS ", ""),
                "Precio Total": band_text(ws, 325, 380).replace("GS ", ""),
                "Impuesto": band_text(ws, 380, 430),
                "Plazo": band_text(ws, 500, 570),
            }

            # Evita falsos positivos en pie de página.
            if row["Código"] and row["Descripción"]:
                rows.append(row)

    return normalize_numeric_columns(ensure_standard_columns(pd.DataFrame(rows)))


def parse_electropar_fields(text):
    flat = clean_text(text)
    fields = {
        "Proveedor": "Electropar",
        "N° Presupuesto/Cotización": "",
        "Fecha": "",
        "Validez": "",
        "Cliente": "",
        "RUC": "",
        "Código de Cliente": "",
        "Condición de venta": "",
        "Vendedor": "",
        "Correo": "",
        "Subtotal": "",
        "IVA": "",
        "Total": "",
    }

    # En este PDF el orden lógico del texto no siempre coincide con el visual:
    # el número aparece antes de "Cotización Nº:" y la fecha inmediatamente después.
    patterns = {
        "N° Presupuesto/Cotización": r"Página\s+1\s+de\s+\d+\s+(\d+)\s+Cotización\s+N[º°o]?\s*:",
        "Fecha": r"Cotización\s+N[º°o]?\s*:\s*(\d{2}/\d{2}/\d{4})\s+Fecha de Cotización",
        "Validez": r"Validez de la oferta\s*:\s*(\d{2}/\d{2}/\d{4})",
        "RUC": r"RUC\s*:\s*([0-9-]+)",
        "Código de Cliente": r"Código de Cliente\s*:\s*([A-Z0-9-]+)",
        "Condición de venta": r"Cond\.\s*de Venta\s*:\s*(.+?)(?=\s+OC del Cliente|\s+Rep\. de Ventas|$)",
        "Correo": r"([\w.\-+]+@electropar\.com\.py)",
        "Total": r"TOTAL\s+GS\s+([0-9.]+)",
    }

    for key, pat in patterns.items():
        m = re.search(pat, flat, re.I)
        if m:
            fields[key] = clean_text(m.group(1))

    m = re.search(r"Señor\(es\)\s*:\s*(.+?)\s+RUC\s*:", flat, re.I)
    if m:
        fields["Cliente"] = clean_text(m.group(1))

    # El extractor nativo coloca "E-mail:" antes del nombre del vendedor.
    m = re.search(
        r"Rep\.\s*de Ventas\s*:\s*E-mail\s*:\s*(.+?)\s+[\w.\-+]+@electropar\.com\.py",
        flat, re.I
    )
    if m:
        fields["Vendedor"] = clean_text(m.group(1))

    # Totales en esta plantilla: IVA 10% y total.
    iva_matches = re.findall(r"\bGS\s+([0-9.]+)", flat)
    # El texto nativo suele contener al final: GS 1.847.076 / TOTAL / GS 0 GS 167.916
    m = re.search(r"GS\s+0\s+GS\s+([0-9.]+)", flat)
    if m:
        fields["IVA"] = clean_text(m.group(1))
    fields["Subtotal"] = fields["Total"]

    return fields


# ============================================================
# PARSER PDF NATIVO: COMPAÑÍA COMERCIAL DEL PARAGUAY
# ============================================================
def parse_ccp_rows(data):
    doc = fitz.open(stream=data, filetype="pdf")
    rows = []

    for page in doc:
        words = page.get_text("words")
        starters = []
        for w in words:
            x0, y0, x1, y1, txt = w[:5]
            if 45 <= x0 < 60 and y0 > 250 and re.fullmatch(r"\d{1,3}", txt):
                starters.append((int(txt), y0))

        starters = sorted(set(starters), key=lambda x: x[1])

        for i, (item, sy) in enumerate(starters):
            ey = starters[i + 1][1] - 2 if i + 1 < len(starters) else min(page.rect.height - 80, sy + 55)
            ws = words_by_band(page, sy - 3, ey)

            row = {
                "Item": item,
                "Código": clean_code(band_text(ws, 60, 110)),
                "Descripción": band_text(ws, 110, 350),
                "Referencia": "",
                "Marca": "",
                "Cantidad": band_text(ws, 350, 400),
                "Precio Unitario": band_text(ws, 400, 465),
                "Precio Total": band_text(ws, 465, 515),
                "Impuesto": "IVA 10%",
                "Plazo": "",
            }

            # En este formato la marca está integrada al inicio de la descripción:
            # FINDER-, LOVATO-, etc. La preservamos también como campo Marca.
            m = re.match(r"([A-ZÁÉÍÓÚÑ0-9]+)-", row["Descripción"])
            if m:
                row["Marca"] = m.group(1)

            if row["Código"] and row["Descripción"]:
                rows.append(row)

    return normalize_numeric_columns(ensure_standard_columns(pd.DataFrame(rows)))


def parse_ccp_fields(text):
    flat = clean_text(text)
    fields = {
        "Proveedor": "Compañía Comercial del Paraguay",
        "N° Presupuesto/Cotización": "",
        "Fecha": "",
        "Validez": "",
        "Cliente": "",
        "RUC": "",
        "Código de Cliente": "",
        "Condición de venta": "",
        "Vendedor": "",
        "Correo": "",
        "Subtotal": "",
        "IVA": "",
        "Total": "",
    }

    m = re.search(r"PRESUP\.\s*Chile\s*Nro\.\s*:?\s*(\d+)", flat, re.I)
    if m:
        fields["N° Presupuesto/Cotización"] = m.group(1)

    # Ej.: Emisión 13 de Agosto de 2026 Válido hasta ...
    months = {
        "enero":"01","febrero":"02","marzo":"03","abril":"04","mayo":"05","junio":"06",
        "julio":"07","agosto":"08","septiembre":"09","setiembre":"09","octubre":"10",
        "noviembre":"11","diciembre":"12"
    }

    def convert_spanish_date(s):
        m = re.search(r"(\d{1,2})\s+de\s+([A-Za-zÁÉÍÓÚáéíóú]+)\s+de\s+(\d{4})", s, re.I)
        if not m:
            return clean_text(s)
        day, month, year = m.groups()
        mm = months.get(month.lower(), "")
        return f"{int(day):02d}/{mm}/{year}" if mm else clean_text(s)

    m = re.search(r"Emisión\s+(.+?)\s+Válido hasta", flat, re.I)
    if m:
        fields["Fecha"] = convert_spanish_date(m.group(1))

    m = re.search(r"Válido hasta\s+(.+?)\s+Cond\.\s*de Vta\.", flat, re.I)
    if m:
        fields["Validez"] = convert_spanish_date(m.group(1))

    m = re.search(r"Cond\.\s*de Vta\.\s*(.+?)\s+Sr\(es\)", flat, re.I)
    if m:
        fields["Condición de venta"] = clean_text(m.group(1))

    m = re.search(r"Sr\(es\)\s+([A-Z0-9]+)\s*-\s*(.+?)\s+RUC\s+([0-9-]+)", flat, re.I)
    if m:
        fields["Código de Cliente"] = m.group(1)
        fields["Cliente"] = clean_text(m.group(2))
        fields["RUC"] = m.group(3)

    m = re.search(r"Vendedor\s+(.+?)\s+Ln\s+Código", flat, re.I)
    if m:
        fields["Vendedor"] = clean_text(m.group(1))

    for key, pat in {
        "Subtotal": r"Total Bruto\s+([0-9.]+)",
        "IVA": r"\bIVA\s+([0-9.]+)",
        "Total": r"Total General\s+([0-9.]+)",
    }.items():
        m = re.search(pat, flat, re.I)
        if m:
            fields[key] = m.group(1)

    return fields


# ============================================================
# OCR / TABLA GENÉRICA PARA ELECTRO SYSTEM Y ESCANEADOS
# ============================================================
def preprocess(img_bgr):
    h, w = img_bgr.shape[:2]
    if w < 1800:
        scale = 1800 / w
        img_bgr = cv2.resize(img_bgr, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (3, 3), 0)
    bw = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                               cv2.THRESH_BINARY, 31, 11)
    return img_bgr, gray, bw


def ocr_image(img, psm=6):
    for lang in ("spa+eng", "spa", "eng"):
        try:
            txt = pytesseract.image_to_string(img, lang=lang, config=f"--psm {psm}")
            if txt.strip():
                return txt
        except Exception:
            pass
    try:
        return pytesseract.image_to_string(img, config=f"--psm {psm}")
    except Exception:
        return ""


def cluster(values, tolerance=10):
    values = sorted(int(v) for v in values)
    if not values:
        return []
    groups = [[values[0]]]
    for v in values[1:]:
        if v - groups[-1][-1] <= tolerance:
            groups[-1].append(v)
        else:
            groups.append([v])
    return [int(round(sum(g) / len(g))) for g in groups]


def detect_table(binary):
    inv = 255 - binary
    hk = cv2.getStructuringElement(cv2.MORPH_RECT, (max(30, binary.shape[1] // 25), 1))
    vk = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(30, binary.shape[0] // 25)))
    hs = cv2.morphologyEx(inv, cv2.MORPH_OPEN, hk)
    vs = cv2.morphologyEx(inv, cv2.MORPH_OPEN, vk)
    grid = cv2.add(hs, vs)
    contours, _ = cv2.findContours(grid, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    H, W = binary.shape
    candidates = []
    for c in contours:
        x, y, w, h = cv2.boundingRect(c)
        if w > .42 * W and h > .055 * H and w * h > .035 * W * H:
            candidates.append((x, y, w, h))
    return max(candidates, key=lambda r: r[2] * r[3]) if candidates else None


def extract_grid(img_bgr, rect, binary):
    x, y, w, h = rect
    roi = binary[y:y+h, x:x+w]
    inv = 255 - roi
    hk = cv2.getStructuringElement(cv2.MORPH_RECT, (max(20, w // 20), 1))
    vk = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(20, h // 20)))
    hs = cv2.morphologyEx(inv, cv2.MORPH_OPEN, hk)
    vs = cv2.morphologyEx(inv, cv2.MORPH_OPEN, vk)
    hp = np.sum(hs > 0, axis=1)
    vp = np.sum(vs > 0, axis=0)
    ys = cluster(np.where(hp > .35 * w)[0], 10)
    xs = cluster(np.where(vp > .35 * h)[0], 10)
    xs = cluster([0] + xs + [w-1], 12)
    ys = cluster([0] + ys + [h-1], 12)
    xs = [v for i, v in enumerate(xs) if i == 0 or v - xs[i-1] > 15]
    ys = [v for i, v in enumerate(ys) if i == 0 or v - ys[i-1] > 12]
    if len(xs) < 3 or len(ys) < 3:
        return pd.DataFrame()

    rows = []
    for r in range(len(ys)-1):
        row = []
        for c in range(len(xs)-1):
            x1, x2 = xs[c], xs[c+1]
            y1, y2 = ys[r], ys[r+1]
            px = max(4, int((x2-x1)*.04))
            py = max(4, int((y2-y1)*.08))
            cell = img_bgr[y+y1+py:y+y2-py, x+x1+px:x+x2-px]
            if cell.size == 0:
                row.append("")
                continue
            g = cv2.cvtColor(cell, cv2.COLOR_BGR2GRAY)
            g = cv2.resize(g, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
            row.append(clean_text(ocr_image(g, 6)))
        rows.append(row)

    return pd.DataFrame(rows)


def electro_system_table_from_grid(grid):
    if grid is None or grid.empty:
        return empty_standard_df()

    # Busca una fila de encabezado con 6 columnas aproximadamente.
    header_idx = None
    for i, row in grid.iterrows():
        txt = " ".join(str(v) for v in row.values).lower()
        if "código" in txt and "descrip" in txt and ("precio" in txt or "cantidad" in txt):
            header_idx = i
            break

    data = grid.iloc[(header_idx + 1 if header_idx is not None else 0):].copy()
    data = data.replace(r"^\s*$", np.nan, regex=True).dropna(how="all").fillna("")
    if data.empty:
        return empty_standard_df()

    # La plantilla Electro System tiene 6 columnas.
    if data.shape[1] >= 6:
        data = data.iloc[:, :6]
        data.columns = ["Item", "Código", "Descripción", "Precio Unitario", "Cantidad", "Precio Total"]
        data["Referencia"] = ""
        data["Marca"] = data["Descripción"].map(
            lambda x: (re.search(r"-\s*([A-Z][A-Z0-9]+)\s*$", str(x)) or [None, ""])[1]
            if str(x).strip() else ""
        )
        data["Impuesto"] = ""
        data["Plazo"] = ""
        # Correcciones OCR conservadoras.
        data["Descripción"] = data["Descripción"].astype(str).str.replace(
            r"\bATAKOM\b", "DATAKOM", regex=True
        )
        return normalize_numeric_columns(ensure_standard_columns(data))

    return empty_standard_df()


def parse_electro_system_fields(text):
    flat = clean_text(text)
    fields = {
        "Proveedor": "Electro System",
        "N° Presupuesto/Cotización": "",
        "Fecha": "",
        "Validez": "",
        "Cliente": "",
        "RUC": "",
        "Código de Cliente": "",
        "Condición de venta": "",
        "Vendedor": "",
        "Correo": "",
        "Subtotal": "",
        "IVA": "",
        "Total": "",
    }

    pats = {
        "N° Presupuesto/Cotización": r"Presupuesto\s*N[°ºo]?\s*:\s*([A-Z0-9-]+)",
        "Fecha": r"Fecha\s*:\s*(\d{1,2}/\d{1,2}/\d{4})",
        "RUC": r"RUC\s*:\s*([0-9-]+)",
        "Código de Cliente": r"C[oó]digo de Cliente\s*:\s*([A-Z0-9-]+)",
        "Condición de venta": r"Condici[oó]n Venta\s*:\s*(.+?)(?=\s+Plazo de Entrega|$)",
        "Vendedor": r"Vendedor\s*:\s*(.+?)(?=\s+Correo\s*:|$)",
        "Correo": r"Correo\s*:\s*([^\s]+)",
        "Subtotal": r"SUB TOTAL\s*:\s*([0-9.]+)",
        "IVA": r"I\.?V\.?A\.?\s*10%\s*:\s*([0-9.]+)",
    }

    for key, pat in pats.items():
        m = re.search(pat, flat, re.I)
        if m:
            fields[key] = clean_text(m.group(1))

    m = re.search(r"Señor\(es\)\s*:\s*(.+?)(?=\s+RUC\s*:|\s+Direccion|\s+Dirección|$)", flat, re.I)
    if m:
        client = clean_text(m.group(1))
        client = re.sub(r"\s+(ee|e|oo)$", "", client, flags=re.I)
        fields["Cliente"] = client

    m = re.search(r"Validez de la oferta\s*:\s*(\d+)\s*d[ií]as", flat, re.I)
    if m:
        fields["Validez"] = f"{m.group(1)} días"

    return fields


def load_images(data, filename):
    if Path(filename).suffix.lower() == ".pdf":
        doc = fitz.open(stream=data, filetype="pdf")
        pages = []
        for page in doc:
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            arr = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, 3)
            pages.append(cv2.cvtColor(arr, cv2.COLOR_RGB2BGR))
        return pages
    arr = np.frombuffer(data, np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    return [img] if img is not None else []


def process_ocr_generic(data, filename, provider_hint="Genérico / escaneado"):
    pages = load_images(data, filename)
    raw_all = []
    tables = []
    previews = []

    for img in pages:
        work, gray, bw = preprocess(img)
        raw = ocr_image(gray, 6)
        raw_all.append(raw)
        rect = detect_table(bw)
        if rect:
            grid = extract_grid(work, rect, bw)
            if not grid.empty:
                tables.append(grid)
        previews.append(work)

    raw_text = "\n".join(raw_all)
    provider = provider_hint
    if "electro system" in raw_text.lower() or "presupuesto de venta" in raw_text.lower():
        provider = "Electro System"

    if provider == "Electro System":
        table = electro_system_table_from_grid(tables[0] if tables else pd.DataFrame())
        fields = parse_electro_system_fields(raw_text)
    else:
        table = empty_standard_df()
        fields = {"Proveedor": provider, "N° Presupuesto/Cotización": "", "Fecha": "",
                  "Validez": "", "Cliente": "", "RUC": "", "Código de Cliente": "",
                  "Condición de venta": "", "Vendedor": "", "Correo": "",
                  "Subtotal": "", "IVA": "", "Total": ""}

    return provider, fields, table, raw_text, previews


# ============================================================
# PROCESAMIENTO GENERAL
# ============================================================
def process_document(data, filename):
    suffix = Path(filename).suffix.lower()
    native_text = ""

    if suffix == ".pdf":
        try:
            native_text = native_pdf_text(data)
        except Exception:
            native_text = ""

    provider = detect_provider(native_text, filename)

    if provider == "Electropar" and len(native_text.strip()) > 80:
        fields = parse_electropar_fields(native_text)
        table = parse_electropar_rows(data)
        previews = load_images(data, filename)
        return provider, fields, table, native_text, previews, "PDF nativo"

    if provider == "Compañía Comercial del Paraguay" and len(native_text.strip()) > 80:
        fields = parse_ccp_fields(native_text)
        table = parse_ccp_rows(data)
        previews = load_images(data, filename)
        return provider, fields, table, native_text, previews, "PDF nativo"

    provider, fields, table, raw, previews = process_ocr_generic(data, filename, provider)
    return provider, fields, table, raw, previews, "OCR"


def validation_report(df):
    if df.empty:
        return pd.DataFrame(columns=["Item", "Estado", "Diferencia"])

    rows = []
    for _, r in df.iterrows():
        qty = parse_py_amount(r.get("Cantidad", ""))
        unit = parse_py_amount(r.get("Precio Unitario", ""))
        total = parse_py_amount(r.get("Precio Total", ""))
        if qty is None or unit is None or total is None:
            rows.append({"Item": r.get("Item", ""), "Estado": "Sin validar", "Diferencia": ""})
            continue
        diff = round(qty * unit - total, 2)
        ok = abs(diff) <= max(1.0, abs(total) * 0.001)
        rows.append({"Item": r.get("Item", ""), "Estado": "OK" if ok else "REVISAR", "Diferencia": diff})
    return pd.DataFrame(rows)


def make_excel(fields, table_df, validation_df, source_name):
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        pd.DataFrame({"Campo": list(fields.keys()), "Valor": list(fields.values())}).to_excel(
            writer, index=False, sheet_name="Datos"
        )
        table_df.to_excel(writer, index=False, sheet_name="Productos")
        validation_df.to_excel(writer, index=False, sheet_name="Validación")

    out.seek(0)
    wb = load_workbook(out)
    dark = PatternFill("solid", fgColor="1F2937")
    white_bold = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")

    for ws in wb.worksheets:
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = dark
                cell.font = white_bold
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb["Datos"].column_dimensions["A"].width = 30
    wb["Datos"].column_dimensions["B"].width = 65

    wp = wb["Productos"]
    widths = {
        "A": 9, "B": 16, "C": 55, "D": 20, "E": 16,
        "F": 12, "G": 18, "H": 18, "I": 14, "J": 22
    }
    for c, width in widths.items():
        wp.column_dimensions[c].width = width

    # Formato numérico real en Excel.
    header_map = {wp.cell(1, c).value: c for c in range(1, wp.max_column + 1)}
    for name in ["Cantidad", "Precio Unitario", "Precio Total"]:
        col = header_map.get(name)
        if col:
            for r in range(2, wp.max_row + 1):
                cell = wp.cell(r, col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00' if name == "Cantidad" else '#,##0'

    info = wb.create_sheet("Info")
    info.append(["Archivo original", source_name])
    info.append(["Aplicación", f"{APP_NAME} v{APP_VERSION}"])
    info.append(["Aviso", "Revisar datos OCR antes de usar comercial o contablemente."])
    info.column_dimensions["A"].width = 25
    info.column_dimensions["B"].width = 85

    final = io.BytesIO()
    wb.save(final)
    return final.getvalue()


# ============================================================
# ESTADO
# ============================================================
defaults = {
    "processed": False,
    "provider": "",
    "fields": {},
    "table": empty_standard_df(),
    "raw": "",
    "previews": [],
    "source_mode": "",
    "filename": "",
}
for key, val in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = val


# ============================================================
# INTERFAZ
# ============================================================
with st.sidebar:
    st.markdown(f"## 📄 {APP_NAME}")
    st.caption(f"Versión {APP_VERSION}")
    st.markdown("---")
    st.markdown("**Formatos optimizados**")
    st.write("✅ Electro System")
    st.write("✅ Electropar")
    st.write("✅ Compañía Comercial del Paraguay")
    st.markdown("---")
    st.caption("PDF · JPG · JPEG · PNG")

st.markdown(f'<div class="hero-title">📄 {APP_NAME}</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="hero-subtitle">Lee presupuestos, identifica el proveedor, extrae productos y genera Excel editable.</div>',
    unsafe_allow_html=True
)

uploaded = st.file_uploader("Subí un presupuesto", type=["pdf", "jpg", "jpeg", "png"])

if uploaded is None:
    c1, c2, c3 = st.columns(3)
    c1.markdown("### 1 · Cargar")
    c1.write("PDF o imagen del proveedor.")
    c2.markdown("### 2 · Detectar")
    c2.write("Identifica formato y extrae productos.")
    c3.markdown("### 3 · Validar")
    c3.write("Corregí y descargá Excel.")
    st.stop()

if st.button("🔎 Procesar presupuesto", type="primary", use_container_width=True):
    with st.spinner("Detectando proveedor y extrayendo datos..."):
        try:
            provider, fields, table, raw, previews, source_mode = process_document(
                uploaded.getvalue(), uploaded.name
            )
            st.session_state.processed = True
            st.session_state.provider = provider
            st.session_state.fields = fields
            st.session_state.table = table
            st.session_state.raw = raw
            st.session_state.previews = previews
            st.session_state.source_mode = source_mode
            st.session_state.filename = uploaded.name
        except Exception as exc:
            st.error(f"No se pudo procesar el documento: {exc}")
            st.stop()

if not st.session_state.processed:
    st.info("Presioná **Procesar presupuesto**.")
    st.stop()

st.markdown(
    f'Proveedor detectado: <span class="provider">{st.session_state.provider}</span> '
    f'· Lectura: **{st.session_state.source_mode}**',
    unsafe_allow_html=True
)

validation = validation_report(st.session_state.table)

m1, m2, m3, m4 = st.columns(4)
m1.metric("Productos", len(st.session_state.table))
m2.metric("Campos", sum(bool(str(v).strip()) for v in st.session_state.fields.values()))
m3.metric("Validaciones OK", int((validation["Estado"] == "OK").sum()) if not validation.empty else 0)
m4.metric("A revisar", int((validation["Estado"] == "REVISAR").sum()) if not validation.empty else 0)

tab1, tab2, tab3, tab4, tab5 = st.tabs(
    ["📋 Datos", "🧾 Productos", "✅ Validación", "🖼️ Vista previa", "🔍 Texto detectado"]
)

with tab1:
    fdf = pd.DataFrame({
        "Campo": list(st.session_state.fields.keys()),
        "Valor": list(st.session_state.fields.values())
    })
    edited = st.data_editor(
        fdf, use_container_width=True, hide_index=True, num_rows="fixed",
        column_config={
            "Campo": st.column_config.TextColumn(disabled=True),
            "Valor": st.column_config.TextColumn(width="large")
        },
        key="fields_editor"
    )
    st.session_state.fields = dict(zip(edited["Campo"], edited["Valor"]))

with tab2:
    st.caption("Los precios y cantidades se intentan guardar como números reales para Excel.")
    st.session_state.table = st.data_editor(
        st.session_state.table,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        key="products_editor"
    )

with tab3:
    validation = validation_report(st.session_state.table)
    st.dataframe(validation, use_container_width=True, hide_index=True)
    if not validation.empty and (validation["Estado"] == "REVISAR").any():
        st.warning("Hay filas donde Cantidad × Precio Unitario no coincide con Precio Total.")
    elif not validation.empty:
        st.success("Las filas con datos numéricos completos pasaron la validación.")

with tab4:
    previews = st.session_state.previews
    if previews:
        p = st.selectbox("Página", list(range(1, len(previews)+1)), format_func=lambda x: f"Página {x}")
        rgb = cv2.cvtColor(previews[p-1], cv2.COLOR_BGR2RGB)
        st.image(rgb, use_container_width=True)

with tab5:
    st.text_area("Texto extraído / OCR", st.session_state.raw, height=430)

validation = validation_report(st.session_state.table)
excel = make_excel(
    st.session_state.fields,
    st.session_state.table,
    validation,
    st.session_state.filename
)

st.markdown("### 📥 Exportar")
st.download_button(
    "⬇️ Descargar Excel",
    data=excel,
    file_name=f"{Path(st.session_state.filename).stem}_extraido.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
    use_container_width=True
)
st.caption("Revisá los campos antes de utilizar la información comercial o contablemente.")
